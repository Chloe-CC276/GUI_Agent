"""Generate deterministic 100-row Excel fixtures for the four demo systems."""

from __future__ import annotations

from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


MVP_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = MVP_ROOT / "test_data"
BASE_TIME = datetime(2026, 1, 5, 9, 0, 0)
MONEY = Decimal("0.01")

DEPARTMENTS = ["研发部", "行政部", "综合管理部", "市场部", "财务部", "生产部", "安全环保部", "信息中心"]
APPLICANTS = ["张明", "李华", "王芳", "赵强", "陈晨", "刘洋", "孙悦", "周宁", "吴桐", "郑凯"]
BUDGET_PROJECTS = [
    ("研发办公设备预算", "BUD-RD-2026"),
    ("行政办公耗材", "BUD-ADM-2026"),
    ("设施更新预算", "BUD-FAC-2026"),
    ("市场推广设备", "BUD-MKT-2026"),
    ("信息化基础设施", "BUD-IT-2026"),
    ("安全生产专项", "BUD-SAFE-2026"),
    ("培训设施预算", "BUD-HR-2026"),
    ("财务系统配套", "BUD-FIN-2026"),
]
COST_CENTERS = ["CC-RD", "CC-ADM", "CC-FAC", "CC-MKT", "CC-IT", "CC-SAFE", "CC-HR", "CC-FIN"]
URGENCY_LEVELS = ["NORMAL", "URGENT", "CRITICAL"]
REQUESTED_METHODS = ["online", "inquiry", "bidding", "centralized", "single", "framework"]
PURCHASE_REASONS = [
    "岗位扩编后需补充办公设备，保障日常研发与协作效率。",
    "库存耗材不足，需及时补货避免影响行政办公运转。",
    "现有设备老化故障率升高，需更新以满足会议与演示需求。",
    "业务外勤增多，需配备移动终端支持现场拜访与资料查阅。",
    "机房扩容配套网络设备，保障信息系统稳定运行。",
    "安全生产检查要求补齐防护与检测器材。",
    "新员工培训场地缺少投影与音响设备，影响集中授课效果。",
    "财务月结高峰需增加打印与扫描能力，缩短单据处理时间。",
]
REMARKS = [
    "优先选择国产品牌，便于维保。",
    "需与现有资产台账编码对齐。",
    "到货后请通知申请人现场验收。",
    "可分两批到货，第一批先满足核心岗位。",
    "请采购员关注质保期不少于三年。",
]
MATERIAL_BASES = [
    ("商务笔记本电脑", "14英寸/32GB/1TB", "台", Decimal("8800")),
    ("27英寸显示器", "4K IPS", "台", Decimal("3200")),
    ("A4复印纸", "80g 500张/包", "包", Decimal("28.50")),
    ("视频会议摄像头", "4K USB", "套", Decimal("6200")),
    ("办公桌", "1600×800mm", "张", Decimal("1850")),
    ("人体工学椅", "网布/可调扶手", "把", Decimal("1250")),
    ("千兆交换机", "24口/可管理", "台", Decimal("2680")),
    ("移动固态硬盘", "2TB/USB-C", "块", Decimal("980")),
    ("激光打印机", "双面/网络打印", "台", Decimal("4600")),
    ("投影仪", "5000流明/1080P", "台", Decimal("7900")),
]
SUPPLIER_BASES = [
    ("华创科技有限公司", "offline_inquiry;framework;direct"),
    ("联信办公设备公司", "offline_inquiry;mall;other"),
    ("中建物资供应中心", "framework;direct;other"),
    ("云商城自营店", "mall;direct"),
    ("华北电子商贸", "offline_inquiry;mall"),
    ("鼎盛办公耗材", "mall;other"),
    ("远航设备工程", "framework;direct"),
    ("星河信息技术", "offline_inquiry;framework;mall"),
    ("国联集采中心", "framework;direct;other"),
    ("优采商城旗舰店", "mall;direct;other"),
]
AWARD_SOURCE_LABELS = {
    "offline_inquiry": "线下询比价",
    "framework": "框架协议",
    "mall": "商城",
    "direct": "直接采购",
    "other": "其他",
}


def money(value: Decimal) -> Decimal:
    return value.quantize(MONEY, rounding=ROUND_HALF_UP)


def build_materials() -> list[dict]:
    rows: list[dict] = []
    for index in range(1, 101):
        name, specification, unit, base_price = MATERIAL_BASES[(index - 1) % len(MATERIAL_BASES)]
        variant = (index - 1) // len(MATERIAL_BASES) + 1
        rows.append(
            {
                "material_code": f"ERP-MAT-{index:04d}",
                "material_name": f"{name}-{variant:02d}",
                "specification": f"{specification}/V{variant}",
                "unit": unit,
                "standard_price": money(base_price * (Decimal("1") + Decimal(variant - 1) * Decimal("0.03"))),
                "status": "active" if index <= 90 else "inactive",
            }
        )
    return rows


def build_suppliers() -> list[dict]:
    rows: list[dict] = []
    for index in range(1, 31):
        name, sources = SUPPLIER_BASES[(index - 1) % len(SUPPLIER_BASES)]
        variant = (index - 1) // len(SUPPLIER_BASES) + 1
        rows.append(
            {
                "supplier_code": f"SUP-{index:03d}",
                "supplier_name": f"{name}-{variant:02d}" if variant > 1 else name,
                "status": "active" if index <= 28 else "inactive",
                "award_sources": sources,
                "award_source_names": ";".join(
                    AWARD_SOURCE_LABELS[code] for code in sources.split(";") if code in AWARD_SOURCE_LABELS
                ),
            }
        )
    return rows


def oa_status(index: int) -> str:
    if index <= 70:
        return "APPROVED"
    if index <= 85:
        return "IN_APPROVAL"
    if index <= 95:
        return "DRAFT"
    return "REJECTED"


def build_oa(materials: Sequence[dict]) -> list[dict]:
    rows: list[dict] = []
    for index in range(1, 101):
        material = materials[(index - 1) % 90]
        quantity = Decimal((index % 8) + 1)
        estimated_price = money(material["standard_price"] * Decimal("1.05"))
        department = DEPARTMENTS[(index - 1) % len(DEPARTMENTS)]
        budget_name, budget_code = BUDGET_PROJECTS[(index - 1) % len(BUDGET_PROJECTS)]
        status = oa_status(index)
        # Required for submit: always populated.
        purchase_reason = PURCHASE_REASONS[(index - 1) % len(PURCHASE_REASONS)]
        urgency = URGENCY_LEVELS[(index - 1) % len(URGENCY_LEVELS)]

        # Optional fields: intentionally sparse to reflect real-world diversity.
        # - budget_project_code: ~70% filled
        # - cost_center_code: ~60% filled
        # - requested_method: ~50% filled
        # - expected_completion_date: ~55% filled
        # - remark: ~40% filled
        budget_project_code = budget_code if index % 10 not in {3, 7, 9} else None
        cost_center_code = (
            COST_CENTERS[(index - 1) % len(COST_CENTERS)] if index % 5 != 0 else None
        )
        requested_method = (
            REQUESTED_METHODS[(index - 1) % len(REQUESTED_METHODS)]
            if index % 2 == 0
            else None
        )
        expected_completion_date = (
            (BASE_TIME + timedelta(days=30 + (index % 45))).date()
            if index % 9 not in {1, 4, 8}
            else None
        )
        remark = REMARKS[(index - 1) % len(REMARKS)] if index % 5 in {1, 2} else None

        rows.append(
            {
                "application_no": f"OA-TEST-{index:04d}",
                "title": f"{department}{material['material_name']}采购",
                "applicant": APPLICANTS[(index - 1) % len(APPLICANTS)],
                "department": department,
                "status": status,
                "total_budget": money(quantity * estimated_price),
                "budget_project_name": budget_name,
                "budget_project_code": budget_project_code,
                "cost_center_code": cost_center_code,
                "requested_method": requested_method,
                "purchase_reason": purchase_reason,
                "urgency_level": urgency,
                "expected_completion_date": expected_completion_date,
                "remark": remark,
                "created_at": BASE_TIME + timedelta(days=index - 1),
                "item_name": material["material_name"],
                "specification": material["specification"],
                "quantity": quantity,
                "estimated_unit_price": estimated_price,
            }
        )
    return rows


def build_procurement(materials: Sequence[dict]) -> list[dict]:
    rows: list[dict] = []
    statuses = ["draft", "validated", "ready", "submitted"]
    for index in range(1, 101):
        oa_index = (index - 1) % 70 + 1
        material = materials[(index - 1) % 90]
        quantity = Decimal((index % 6) + 1)
        authoritative_price = material["standard_price"]
        input_price = authoritative_price + (Decimal("1.00") if index % 10 == 0 else Decimal("0"))
        status = statuses[(index - 1) % len(statuses)]
        created_at = BASE_TIME + timedelta(days=index - 1, hours=2)
        rows.append(
            {
                "request_no": f"PR-TEST-{index:04d}",
                "oa_application_no": f"OA-TEST-{oa_index:04d}",
                "business_key": f"OA-TEST-{oa_index:04d}",
                "status": status,
                "material_code": material["material_code"],
                "material_name": material["material_name"],
                "specification": material["specification"],
                "unit": material["unit"],
                "quantity": quantity,
                "unit_price": money(input_price),
                "line_amount": money(quantity * input_price),
                "expected_server_total": money(quantity * authoritative_price),
                "difference_expected": "yes" if index % 10 == 0 else "no",
                "attachment_file_name": f"采购附件_{index:04d}.pdf",
                "attachment_file_url": f"demo://采购附件_{index:04d}.pdf",
                "created_at": created_at,
                "submitted_at": created_at + timedelta(hours=4) if status == "submitted" else None,
            }
        )
    return rows


def build_agent_tasks(procurement: Sequence[dict]) -> list[dict]:
    operations = ["create", "validate", "prepare-submit", "confirm-submit"]
    rows: list[dict] = []
    for index, request in enumerate(procurement, start=1):
        operation = operations[(index - 1) % len(operations)]
        task_id = f"excel-{operation}-{index:04d}"
        rows.append(
            {
                "task_id": task_id,
                "business_key": request["business_key"],
                "operation": operation,
                "status": "completed",
                "result_request_no": request["request_no"],
                "result_request_id": index,
                "result_json": f'{{"request_id":{index},"request_no":"{request["request_no"]}"}}',
                "idempotency_key": f"{task_id}|{request['business_key']}|{operation}",
                "created_at": request["created_at"] + timedelta(minutes=30),
            }
        )
    return rows


def write_workbook(
    path: Path,
    sheet_name: str,
    table_name: str,
    fields: Sequence[tuple[str, str]],
    records: Iterable[dict],
    notes: Sequence[str],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    field_names = [field for field, _ in fields]
    # OA test workbook uses Chinese headers for human填报; other sheets keep English keys.
    if sheet_name == "OA采购申请":
        worksheet.append([description for _, description in fields])
    else:
        worksheet.append(field_names)
    records_list = list(records)
    for record in records_list:
        worksheet.append([record.get(field) for field in field_names])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 24

    for column_cells in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(len(value) for value in values) + 3, 36)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(cell.value, Decimal):
                cell.number_format = "0.00"

    last_row = max(len(records_list) + 1, 2)
    table = Table(
        displayName=table_name,
        ref=f"A1:{worksheet.cell(last_row, len(fields)).coordinate}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    info = workbook.create_sheet("字段说明")
    info.append(["字段名", "中文说明"])
    for field, description in fields:
        info.append([field, description])
    info.append([])
    info.append(["数据规则", "说明"])
    for index, note in enumerate(notes, start=1):
        info.append([f"R{index}", note])
    info.freeze_panes = "A2"
    info.column_dimensions["A"].width = 30
    info.column_dimensions["B"].width = 80
    for cell in info[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    workbook.save(path)


def verify(path: Path, sheet_name: str, expected: int = 100) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    record_count = worksheet.max_row - 1
    workbook.close()
    if record_count != expected:
        raise RuntimeError(f"{path.name}: expected {expected} records, got {record_count}")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    materials = build_materials()
    suppliers = build_suppliers()
    oa_rows = build_oa(materials)
    procurement_rows = build_procurement(materials)
    task_rows = build_agent_tasks(procurement_rows)

    outputs = [
        (
            OUTPUT_DIR / "OA采购申请_100条.xlsx",
            "OA采购申请",
            "OAApplications",
            [
                ("application_no", "OA申请编号"),
                ("title", "申请标题"),
                ("applicant", "申请人"),
                ("department", "申请部门"),
                ("status", "审批状态"),
                ("total_budget", "预算总额"),
                ("budget_project_name", "预算项目名称"),
                ("budget_project_code", "预算项目编码"),
                ("cost_center_code", "成本中心"),
                ("requested_method", "建议采买方式"),
                ("purchase_reason", "采购原因"),
                ("urgency_level", "紧急程度"),
                ("expected_completion_date", "期望完成日期"),
                ("remark", "备注"),
                ("created_at", "创建时间"),
                ("item_name", "需求物资"),
                ("specification", "规格型号"),
                ("quantity", "数量"),
                ("estimated_unit_price", "预估单价"),
            ],
            oa_rows,
            [
                "共100条申请：70条APPROVED、15条IN_APPROVAL、10条DRAFT、5条REJECTED。",
                "提交必填字段均已生成：title/applicant/department/total_budget/budget_project_name/purchase_reason/urgency_level，以及物资行 quantity/estimated_unit_price（预估单价）。",
                "可选字段按比例留空：预算编码约30%空、成本中心约20%空、采买方式约50%空、期望完成日约33%空、备注约60%空。",
                "每条申请包含一条扁平化物资明细（需求物资/规格型号/数量/预估单价），可对照 OA 填报页录入。",
                "审批状态取值：APPROVED/IN_APPROVAL/DRAFT/REJECTED；紧急程度：NORMAL/URGENT/CRITICAL；采买方式可选 online/inquiry/bidding/centralized/single/framework。",
            ],
        ),
        (
            OUTPUT_DIR / "ERP物料主数据_100条.xlsx",
            "ERP物料主数据",
            "ERPMaterials",
            [
                ("material_code", "ERP唯一物料编码"),
                ("material_name", "标准物料名称"),
                ("specification", "标准规格型号"),
                ("unit", "计量单位"),
                ("standard_price", "ERP标准价格"),
                ("status", "主数据状态：active/inactive"),
            ],
            materials,
            ["共100条物料：90条有效、10条停用。", "采购录入只允许选择active物料编码。"],
        ),
        (
            OUTPUT_DIR / "最终供应商信息.xlsx",
            "最终供应商信息",
            "ERPSuppliers",
            [
                ("supplier_code", "供应商编码"),
                ("supplier_name", "供应商名称"),
                ("status", "主数据状态：active/inactive"),
                ("award_sources", "结果来源编码，多值用分号分隔"),
                ("award_source_names", "结果来源中文名，多值用分号分隔"),
            ],
            suppliers,
            [
                "供应商与结果来源为多对多：一行供应商可对应多个 award_sources。",
                "award_sources 取值：offline_inquiry/framework/mall/direct/other。",
                "导入后可在采购申请核对页下拉选择供应商及其允许的结果来源。",
            ],
        ),
        (
            OUTPUT_DIR / "采购云申请_100条.xlsx",
            "采购云申请",
            "ProcurementRequests",
            [
                ("request_no", "采购申请编号"),
                ("oa_application_no", "来源OA申请编号"),
                ("business_key", "跨系统业务键"),
                ("status", "采购状态"),
                ("material_code", "ERP物料编码"),
                ("material_name", "物料名称快照"),
                ("specification", "规格快照"),
                ("unit", "单位"),
                ("quantity", "数量"),
                ("unit_price", "页面输入单价"),
                ("line_amount", "页面输入行金额"),
                ("expected_server_total", "按ERP标准价计算的预期服务端金额"),
                ("difference_expected", "是否预期产生价格差异记录"),
                ("attachment_file_name", "附件文件名"),
                ("attachment_file_url", "演示附件地址"),
                ("created_at", "创建时间"),
                ("submitted_at", "提交时间"),
            ],
            procurement_rows,
            ["全部引用已通过的OA申请及active ERP物料。", "每第10条故意设置1元单价差异，用于验证差异留痕。"],
        ),
        (
            OUTPUT_DIR / "Agent任务_100条.xlsx",
            "Agent任务",
            "AgentTasks",
            [
                ("task_id", "Agent任务唯一编号"),
                ("business_key", "跨系统业务键"),
                ("operation", "幂等操作名称"),
                ("status", "任务状态"),
                ("result_request_no", "结果采购申请编号"),
                ("result_request_id", "结果采购申请ID"),
                ("result_json", "结构化结果"),
                ("idempotency_key", "task_id+business_key+operation组合键"),
                ("created_at", "任务创建时间"),
            ],
            task_rows,
            ["同一idempotency_key重复执行不得重复建单。", "任务与采购申请、OA通过business_key关联。"],
        ),
    ]

    for path, sheet, table, fields, records, notes in outputs:
        record_list = list(records)
        write_workbook(path, sheet, table, fields, record_list, notes)
        verify(path, sheet, expected=len(record_list))
        print(f"Generated {path} ({len(record_list)} records)")


if __name__ == "__main__":
    main()
