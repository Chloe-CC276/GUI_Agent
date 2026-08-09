"""Import ERP materials from test_data Excel into the demo SQLite master data."""

from __future__ import annotations

import argparse
import os
import sys
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select

MVP_ROOT = Path(__file__).resolve().parents[1]
BACKEND_ROOT = MVP_ROOT / "backend"
DEFAULT_XLSX = MVP_ROOT / "test_data" / "ERP物料主数据_100条.xlsx"
DEFAULT_DB = BACKEND_ROOT / "procurement_demo.db"

sys.path.insert(0, str(BACKEND_ROOT))

from app.db import Database  # noqa: E402
from app.models import ERPMaterial  # noqa: E402
from app.seed import init_database  # noqa: E402


HEADER_ALIASES = {
    "material_code": {"material_code", "物料编码", "erp唯一物料编码", "编码"},
    "material_name": {"material_name", "物料名称", "标准物料名称", "名称"},
    "specification": {"specification", "规格型号", "标准规格型号", "规格"},
    "unit": {"unit", "计量单位", "单位"},
    "standard_price": {"standard_price", "erp标准价格", "标准价格", "单价"},
    "status": {"status", "主数据状态：active/inactive", "主数据状态", "状态"},
}


def _norm(value: object) -> str:
    return str(value or "").strip()


def _map_headers(headers: list[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        name = _norm(header).lower()
        for field, aliases in HEADER_ALIASES.items():
            if name in {alias.lower() for alias in aliases}:
                mapping[field] = index
                break
    missing = [field for field in ("material_code", "material_name", "unit", "standard_price") if field not in mapping]
    if missing:
        raise RuntimeError(f"Excel missing required columns: {missing}; headers={headers}")
    return mapping


def load_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows_iter = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if not rows_iter:
        raise RuntimeError(f"{path} is empty")
    mapping = _map_headers(list(rows_iter[0]))
    records: list[dict] = []
    for raw in rows_iter[1:]:
        if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
            continue
        code = _norm(raw[mapping["material_code"]])
        if not code:
            continue
        status = _norm(raw[mapping["status"]]) if "status" in mapping else "active"
        if status.lower() in {"", "有效", "启用"}:
            status = "active"
        elif status.lower() in {"停用", "无效"}:
            status = "inactive"
        records.append(
            {
                "material_code": code,
                "material_name": _norm(raw[mapping["material_name"]]) or code,
                "specification": _norm(raw[mapping["specification"]]) if "specification" in mapping else "",
                "unit": _norm(raw[mapping["unit"]]) or "个",
                "standard_price": Decimal(str(raw[mapping["standard_price"]] or "0")),
                "status": status if status in {"active", "inactive"} else "active",
            }
        )
    return records


def import_materials(database_url: str, excel_path: Path) -> dict[str, int]:
    records = load_rows(excel_path)
    database = Database(database_url)
    init_database(database)
    inserted = 0
    updated = 0
    with database.session_factory() as session:
        for item in records:
            existing = session.scalar(
                select(ERPMaterial).where(ERPMaterial.material_code == item["material_code"])
            )
            if existing is None:
                session.add(ERPMaterial(**item))
                inserted += 1
            else:
                existing.material_name = item["material_name"]
                existing.specification = item["specification"]
                existing.unit = item["unit"]
                existing.standard_price = item["standard_price"]
                existing.status = item["status"]
                updated += 1
        session.commit()
        total = session.scalar(select(func.count(ERPMaterial.id))) or 0
    return {
        "excel_rows": len(records),
        "inserted": inserted,
        "updated": updated,
        "total_in_db": int(total),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import ERP materials Excel into demo DB")
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"Excel path (default: {DEFAULT_XLSX})",
    )
    parser.add_argument(
        "--database-url",
        default=os.getenv("DATABASE_URL", f"sqlite:///{DEFAULT_DB.as_posix()}"),
        help="SQLAlchemy database URL",
    )
    args = parser.parse_args()
    if not args.excel.exists():
        raise SystemExit(f"Excel not found: {args.excel}")
    result = import_materials(args.database_url, args.excel)
    print(
        f"Imported {result['excel_rows']} rows from {args.excel.name}: "
        f"+{result['inserted']} inserted, {result['updated']} updated; "
        f"erp_materials total={result['total_in_db']}"
    )


if __name__ == "__main__":
    main()
