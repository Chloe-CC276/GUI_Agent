"""Import final suppliers + award-source M2M links into the demo SQLite DB."""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import selectinload

MVP_ROOT = Path(__file__).resolve().parents[1]
BACKEND_ROOT = MVP_ROOT / "backend"
DEFAULT_XLSX = MVP_ROOT / "test_data" / "最终供应商信息.xlsx"
DEFAULT_DB = BACKEND_ROOT / "procurement_demo.db"

sys.path.insert(0, str(BACKEND_ROOT))

from app.db import Database  # noqa: E402
from app.migrations import DEFAULT_AWARD_SOURCES  # noqa: E402
from app.models import AwardSource, ERPSupplier  # noqa: E402
from app.seed import init_database  # noqa: E402
from app.services import normalize_award_source  # noqa: E402

HEADER_ALIASES = {
    "supplier_code": {"supplier_code", "供应商编码", "编码"},
    "supplier_name": {"supplier_name", "供应商名称", "名称"},
    "status": {"status", "主数据状态：active/inactive", "主数据状态", "状态"},
    "award_sources": {
        "award_sources",
        "结果来源编码，多值用分号分隔",
        "结果来源编码",
        "结果来源",
    },
}


def _norm(value: object) -> str:
    return str(value or "").strip()


def _map_headers(headers: list[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        name = _norm(header).lower()
        for field, aliases in HEADER_ALIASES.items():
            if name in {alias.lower() for alias in aliases}:
                mapping[field] = index
                break
    missing = [field for field in ("supplier_code", "supplier_name") if field not in mapping]
    if missing:
        raise RuntimeError(f"Excel missing required columns: {missing}; headers={headers}")
    return mapping


def _split_sources(raw: object) -> list[str]:
    text = _norm(raw)
    if not text:
        return []
    parts = [part.strip() for part in text.replace("，", ";").replace(",", ";").split(";")]
    codes: list[str] = []
    for part in parts:
        code = normalize_award_source(part)
        if code and code not in codes:
            codes.append(code)
    return codes


def load_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows_iter = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if not rows_iter:
        raise RuntimeError(f"{path} is empty")
    mapping = _map_headers(list(rows_iter[0]))
    records: list[dict] = []
    for raw in rows_iter[1:]:
        if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
            continue
        code = _norm(raw[mapping["supplier_code"]])
        if not code:
            continue
        status = _norm(raw[mapping["status"]]) if "status" in mapping else "active"
        if status.lower() in {"", "有效", "启用"}:
            status = "active"
        elif status.lower() in {"停用", "无效"}:
            status = "inactive"
        records.append(
            {
                "supplier_code": code,
                "supplier_name": _norm(raw[mapping["supplier_name"]]) or code,
                "status": status if status in {"active", "inactive"} else "active",
                "award_sources": _split_sources(raw[mapping["award_sources"]])
                if "award_sources" in mapping
                else [],
            }
        )
    return records


def ensure_award_sources(session, codes: set[str]) -> dict[str, AwardSource]:
    by_code = {item.code: item for item in session.scalars(select(AwardSource)).all()}
    name_lookup = dict(DEFAULT_AWARD_SOURCES)
    for code in sorted(codes):
        if code in by_code:
            continue
        source = AwardSource(code=code, name=name_lookup.get(code, code), status="active")
        session.add(source)
        by_code[code] = source
    session.flush()
    return by_code


def import_suppliers(database_url: str, excel_path: Path) -> dict[str, int]:
    records = load_rows(excel_path)
    database = Database(database_url)
    init_database(database)
    inserted = 0
    updated = 0
    all_codes = {code for item in records for code in item["award_sources"]}
    with database.session_factory() as session:
        by_code = ensure_award_sources(session, all_codes)
        for item in records:
            existing = session.scalar(
                select(ERPSupplier)
                .where(ERPSupplier.supplier_code == item["supplier_code"])
                .options(selectinload(ERPSupplier.award_sources))
            )
            sources = [by_code[code] for code in item["award_sources"] if code in by_code]
            if existing is None:
                supplier = ERPSupplier(
                    supplier_code=item["supplier_code"],
                    supplier_name=item["supplier_name"],
                    status=item["status"],
                )
                supplier.award_sources = sources
                session.add(supplier)
                inserted += 1
            else:
                existing.supplier_name = item["supplier_name"]
                existing.status = item["status"]
                existing.award_sources = sources
                updated += 1
        session.commit()
        from sqlalchemy import func

        total = session.scalar(select(func.count(ERPSupplier.id))) or 0
        active = session.scalar(
            select(func.count(ERPSupplier.id)).where(ERPSupplier.status == "active")
        ) or 0
    return {
        "excel_rows": len(records),
        "inserted": inserted,
        "updated": updated,
        "total_in_db": int(total),
        "active_in_db": int(active),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import final suppliers Excel into demo DB")
    parser.add_argument("--excel", type=Path, default=DEFAULT_XLSX)
    parser.add_argument(
        "--database-url",
        default=os.getenv("DATABASE_URL", f"sqlite:///{DEFAULT_DB.as_posix()}"),
    )
    args = parser.parse_args()
    if not args.excel.exists():
        raise SystemExit(f"Excel not found: {args.excel}")
    result = import_suppliers(args.database_url, args.excel)
    print(
        f"Imported {result['excel_rows']} suppliers from {args.excel.name}: "
        f"+{result['inserted']} inserted, {result['updated']} updated; "
        f"total={result['total_in_db']} active={result['active_in_db']}"
    )


if __name__ == "__main__":
    main()
