from __future__ import annotations

import io
import json
import uuid
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session, selectinload

from .errors import AppError
from .models import (
    AgentTask,
    CrossSystemDifference,
    ERPMaterial,
    ExportBatch,
    ImportBatch,
    OAApplication,
    ProcurementRequest,
    ProcurementRequestLine,
    WorkflowEvent,
    utcnow,
)
from .oa_services import normalize_oa_status
from .schemas import ProcurementRequestOut
from .services import (
    _unique_number,
    get_request,
    get_request_by_ref,
    money,
)

BACKEND_ROOT = Path(__file__).resolve().parent.parent
EXPORTS_DIR = BACKEND_ROOT / "exports"
RULES_PATH = Path(__file__).resolve().parent / "config" / "purchase_method_rules.json"

STAGE_DEFINITIONS = [
    ("需求填报", "requirement"),
    ("部门审批", "department_approval"),
    ("预算校验", "budget_check"),
    ("采购寻源", "sourcing"),
    ("供应商报价", "quotation"),
    ("定标下单", "award_order"),
    ("合同签署", "contract"),
    ("到货验收", "receiving"),
    ("财务付款", "payment"),
]

COLUMN_ALIASES = {
    "material_code": {"material_code", "物料编码", "物料代码", "编码"},
    "material_name": {"material_name", "物资名称", "物料名称", "名称"},
    "specification": {"specification", "规格", "规格型号", "型号"},
    "quantity": {"quantity", "数量", "采购数量"},
    "unit": {"unit", "单位", "计量单位"},
    "unit_price": {"unit_price", "含税单价", "单价", "含税价"},
}

DEFAULT_PURCHASE_METHOD_RULES: dict[str, Any] = {
    "version": "v1",
    "currency": "CNY",
    "rules": [
        {"method": "网购", "max_amount": 50000, "min_amount": 0},
        {"method": "比价", "max_amount": 200000, "min_amount": 50000},
        {"method": "招标", "max_amount": 1000000, "min_amount": 200000},
        {"method": "集中采购", "min_amount": 1000000},
    ],
}


def record_workflow_event(
    session: Session,
    *,
    business_key: str,
    event_type: str,
    status: str = "info",
    operator: str | None = None,
    detail: dict | None = None,
    commit: bool = False,
) -> WorkflowEvent:
    event = WorkflowEvent(
        business_key=business_key,
        event_type=event_type,
        status=status,
        operator=operator,
        detail_json=detail or {},
        event_time=utcnow(),
    )
    session.add(event)
    if commit:
        session.commit()
    return event


def get_purchase_method_rules() -> dict[str, Any]:
    if RULES_PATH.exists():
        try:
            return json.loads(RULES_PATH.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            pass
    return json.loads(json.dumps(DEFAULT_PURCHASE_METHOD_RULES))


def suggest_purchase_method(amount: Decimal, rules: dict[str, Any] | None = None) -> str:
    cfg = rules or get_purchase_method_rules()
    value = money(amount)
    with_max: list[dict[str, Any]] = []
    open_ended: dict[str, Any] | None = None
    for rule in cfg.get("rules", []):
        if rule.get("max_amount") is None:
            open_ended = rule
        else:
            with_max.append(rule)
    with_max.sort(key=lambda item: Decimal(str(item["max_amount"])))
    for rule in with_max:
        upper = Decimal(str(rule["max_amount"]))
        lower = Decimal(str(rule.get("min_amount", 0)))
        if lower == 0:
            if value <= upper:
                return str(rule["method"])
        elif value > lower and value <= upper:
            return str(rule["method"])
    if open_ended is not None:
        lower = Decimal(str(open_ended.get("min_amount", 0)))
        if value > lower:
            return str(open_ended["method"])
    return "集中采购"


def ensure_purchase_method(request: ProcurementRequest) -> None:
    rules = get_purchase_method_rules()
    request.rule_version = str(rules.get("version", "v1"))
    request.purchase_method_suggested = suggest_purchase_method(request.total_amount, rules)


def compute_stages(
    oa: OAApplication | None,
    pr: ProcurementRequest | None,
) -> list[dict[str, Any]]:
    oa_status = oa.status if oa else None
    pr_status = pr.status if pr else None
    has_po = bool(pr and pr.po_no)

    if has_po:
        completed_through = 6
        current_index = None
    elif pr and pr_status == "submitted":
        completed_through = 5
        current_index = 6
    elif pr and pr_status == "ready":
        completed_through = 5
        current_index = 6
    elif pr and pr_status == "validated":
        completed_through = 4
        current_index = 5
    elif pr and pr_status == "draft":
        completed_through = 3
        current_index = 4
    elif normalize_oa_status(oa_status) == "APPROVED":
        completed_through = 3
        current_index = 4
    elif normalize_oa_status(oa_status) in {"IN_APPROVAL", "REJECTED"}:
        completed_through = 1
        current_index = 2
    elif normalize_oa_status(oa_status) == "DRAFT" or oa:
        completed_through = 0
        current_index = 1
    else:
        completed_through = 0
        current_index = 1

    stages = []
    for index, (name, code) in enumerate(STAGE_DEFINITIONS, start=1):
        if index <= completed_through:
            status = "completed"
        elif current_index is not None and index == current_index:
            status = "current"
        else:
            status = "pending"
        stages.append(
            {
                "index": index,
                "code": code,
                "name": name,
                "status": status,
                "available": index <= 6,
            }
        )
    return stages


def workbench_summary(session: Session) -> dict[str, Any]:
    pending = (
        session.scalar(
            select(func.count(ProcurementRequest.id)).where(
                ProcurementRequest.status.in_(["draft", "validated", "ready"])
            )
        )
        or 0
    )
    approving = (
        session.scalar(
            select(func.count(OAApplication.id)).where(
                OAApplication.status.in_(["IN_APPROVAL", "pending", "approving"])
            )
        )
        or 0
    )
    now = datetime.now(timezone.utc)
    month_start = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
    month_amount = session.scalar(
        select(func.coalesce(func.sum(ProcurementRequest.total_amount), 0)).where(
            ProcurementRequest.created_at >= month_start
        )
    ) or Decimal("0")
    month_amount = money(Decimal(month_amount))
    budget_total = session.scalar(
        select(func.coalesce(func.sum(OAApplication.total_budget), 0))
    ) or Decimal("0")
    budget_total = Decimal(budget_total)
    budget_rate = (
        float(month_amount / budget_total) if budget_total > 0 else 0.0
    )

    current = session.scalar(
        select(ProcurementRequest)
        .options(
            selectinload(ProcurementRequest.lines),
            selectinload(ProcurementRequest.attachments),
            selectinload(ProcurementRequest.oa_application),
        )
        .order_by(ProcurementRequest.updated_at.desc(), ProcurementRequest.id.desc())
        .limit(1)
    )
    oa = current.oa_application if current else None
    if oa is None and current is None:
        oa = session.scalar(
            select(OAApplication)
            .where(OAApplication.status.in_(["APPROVED", "approved"]))
            .order_by(OAApplication.id)
            .limit(1)
        )

    stages = compute_stages(oa, current)
    events = session.scalars(
        select(WorkflowEvent).order_by(WorkflowEvent.event_time.desc()).limit(20)
    ).all()
    current_pr = None
    if current is not None:
        ensure_purchase_method(current)
        current_pr = ProcurementRequestOut.model_validate(current).model_dump(mode="json")

    return {
        "metrics": {
            "pending": int(pending),
            "approving": int(approving),
            "month_amount": str(month_amount),
            "budget_rate": round(budget_rate, 4),
        },
        "current_pr": current_pr,
        "stages": stages,
        "recent_events": [_event_out(item) for item in events],
    }


def _event_out(event: WorkflowEvent) -> dict[str, Any]:
    return {
        "id": event.id,
        "business_key": event.business_key,
        "event_type": event.event_type,
        "status": event.status,
        "operator": event.operator,
        "detail_json": event.detail_json or {},
        "event_time": event.event_time.isoformat() if event.event_time else None,
    }


def list_workflow_events(
    session: Session, business_key: str | None = None, limit: int = 50
) -> list[dict[str, Any]]:
    statement = select(WorkflowEvent).order_by(WorkflowEvent.event_time.desc()).limit(limit)
    if business_key:
        statement = (
            select(WorkflowEvent)
            .where(WorkflowEvent.business_key == business_key)
            .order_by(WorkflowEvent.event_time.desc())
            .limit(limit)
        )
    return [_event_out(item) for item in session.scalars(statement).all()]


def _normalize_header(value: Any) -> str:
    return str(value or "").strip()


def _map_headers(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        name = _normalize_header(header)
        for field, aliases in COLUMN_ALIASES.items():
            if name in aliases and field not in mapping:
                mapping[field] = index
    return mapping


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).strip().replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def _match_erp_candidates(
    session: Session,
    material_code: str | None,
    material_name: str | None,
    specification: str | None,
) -> tuple[list[dict[str, Any]], Decimal | None, list[str]]:
    errors: list[str] = []
    candidates: list[dict[str, Any]] = []
    confidence: Decimal | None = None

    if material_code:
        material = session.scalar(
            select(ERPMaterial).where(ERPMaterial.material_code == material_code)
        )
        if material is None:
            errors.append("MATERIAL_NOT_FOUND")
        elif material.status != "active":
            errors.append("MATERIAL_INACTIVE")
        else:
            candidates.append(
                {
                    "material_code": material.material_code,
                    "material_name": material.material_name,
                    "specification": material.specification,
                    "unit": material.unit,
                    "standard_price": str(money(material.standard_price)),
                    "status": material.status,
                }
            )
            confidence = Decimal("1.0000")
            return candidates, confidence, errors

    if material_name:
        rows = session.scalars(
            select(ERPMaterial)
            .where(
                ERPMaterial.status == "active",
                or_(
                    ERPMaterial.material_name == material_name,
                    ERPMaterial.material_name.ilike(f"%{material_name}%"),
                ),
            )
            .limit(5)
        ).all()
        for row in rows:
            score = Decimal("0.9000") if row.material_name == material_name else Decimal("0.6500")
            if specification and row.specification == specification:
                score = min(Decimal("0.9900"), score + Decimal("0.0800"))
            candidates.append(
                {
                    "material_code": row.material_code,
                    "material_name": row.material_name,
                    "specification": row.specification,
                    "unit": row.unit,
                    "standard_price": str(money(row.standard_price)),
                    "status": row.status,
                    "match_confidence": str(score),
                }
            )
        if candidates:
            confidence = Decimal(candidates[0].get("match_confidence", "0.6500"))
        else:
            errors.append("NO_ERP_CANDIDATE")
    else:
        errors.append("MISSING_MATERIAL_IDENTITY")
    return candidates, confidence, errors


def preview_import(session: Session, filename: str, content: bytes) -> dict[str, Any]:
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface parse errors to client
        raise AppError(422, "INVALID_XLSX", "Unable to parse Excel workbook", str(exc)) from exc
    sheet = workbook.active
    rows_iter = list(sheet.iter_rows(values_only=True))
    if not rows_iter:
        raise AppError(422, "EMPTY_WORKBOOK", "Excel workbook has no rows")
    headers = [_normalize_header(value) for value in rows_iter[0]]
    mapping = _map_headers(headers)
    required = {"material_name", "quantity"}
    if not required.issubset(mapping) and "material_code" not in mapping:
        raise AppError(
            422,
            "TEMPLATE_MISMATCH",
            "Standard template columns not found",
            {"mapped": mapping, "headers": headers},
        )

    preview_rows: list[dict[str, Any]] = []
    success = 0
    failed = 0
    for row_no, values in enumerate(rows_iter[1:], start=2):
        if values is None or all(cell is None or str(cell).strip() == "" for cell in values):
            continue
        raw = {
            field: values[index] if index < len(values) else None
            for field, index in mapping.items()
        }
        material_code = _normalize_header(raw.get("material_code")) or None
        material_name = _normalize_header(raw.get("material_name")) or None
        specification = _normalize_header(raw.get("specification")) or None
        unit = _normalize_header(raw.get("unit")) or None
        quantity = _to_decimal(raw.get("quantity"))
        unit_price = _to_decimal(raw.get("unit_price"))
        errors: list[str] = []
        if quantity is None or quantity <= 0:
            errors.append("INVALID_QUANTITY")
        candidates, confidence, match_errors = _match_erp_candidates(
            session, material_code, material_name, specification
        )
        errors.extend(match_errors)
        matched = candidates[0] if len(candidates) == 1 and not match_errors else None
        if matched is None and not errors:
            errors.append("AMBIGUOUS_OR_UNMATCHED")
        row_status = "ok" if matched and not errors else "error"
        if row_status == "ok":
            success += 1
        else:
            failed += 1
        preview_rows.append(
            {
                "row_no": row_no,
                "raw": {
                    "material_code": material_code,
                    "material_name": material_name,
                    "specification": specification,
                    "quantity": str(quantity) if quantity is not None else None,
                    "unit": unit,
                    "unit_price": str(unit_price) if unit_price is not None else None,
                },
                "matched": matched,
                "candidates": candidates,
                "match_confidence": str(confidence) if confidence is not None else None,
                "errors": errors,
                "status": row_status,
            }
        )

    batch_id = f"IMP-{uuid.uuid4().hex[:12].upper()}"
    batch = ImportBatch(
        import_batch_id=batch_id,
        filename=filename,
        sheet_name=sheet.title,
        total_rows=len(preview_rows),
        success_rows=success,
        failed_rows=failed,
        status="review",
        preview_json={"rows": preview_rows, "headers": headers, "mapping": mapping},
    )
    session.add(batch)
    session.commit()
    return {
        "import_batch_id": batch_id,
        "filename": filename,
        "sheet_name": sheet.title,
        "total_rows": len(preview_rows),
        "success_rows": success,
        "failed_rows": failed,
        "rows": preview_rows,
    }


def confirm_import(
    session: Session,
    *,
    import_batch_id: str,
    task_id: str,
    business_key: str,
    pr_no: str | None = None,
    oa_application_id: int | None = None,
    row_nos: list[int] | None = None,
) -> dict[str, Any]:
    batch = session.scalar(
        select(ImportBatch).where(ImportBatch.import_batch_id == import_batch_id)
    )
    if batch is None:
        raise AppError(404, "IMPORT_BATCH_NOT_FOUND", "Import batch not found")
    if batch.status == "confirmed" and batch.preview_json.get("confirmed_pr_no"):
        request = get_request_by_ref(session, batch.preview_json["confirmed_pr_no"])
        return {
            "pr_no": request.request_no,
            "import_batch_id": batch.import_batch_id,
            "written_rows": batch.success_rows,
            "request": ProcurementRequestOut.model_validate(request).model_dump(mode="json"),
            "idempotent_replay": True,
        }

    rows = batch.preview_json.get("rows", [])
    selected = [
        row
        for row in rows
        if row_nos is None or row["row_no"] in row_nos
    ]
    writable = [row for row in selected if row.get("status") == "ok" and row.get("matched")]
    skipped = [row for row in selected if row not in writable]
    if not writable:
        raise AppError(
            422,
            "NO_VALID_IMPORT_ROWS",
            "No valid rows available for import; invalid materials cannot be written",
            {"skipped": skipped},
        )

    if pr_no:
        request = get_request_by_ref(session, pr_no)
        if request.status != "draft":
            raise AppError(409, "NOT_EDITABLE", "Only draft requests can accept imports")
    else:
        oa_id = oa_application_id
        if oa_id is None:
            oa = session.scalar(
                select(OAApplication)
                .where(OAApplication.status.in_(["APPROVED", "approved"]))
                .order_by(OAApplication.id)
                .limit(1)
            )
            if oa is None:
                raise AppError(404, "OA_NOT_FOUND", "No approved OA available for import")
            oa_id = oa.id
        else:
            oa = session.get(OAApplication, oa_id)
            if oa is None:
                raise AppError(404, "OA_NOT_FOUND", "OA application not found")
        request = ProcurementRequest(
            request_no=_unique_number("PR-2026-"),
            oa_application_id=oa_id,
            oa_apply_no=oa.application_no,
            oa_title=oa.title,
            oa_applicant=oa.applicant,
            oa_department=oa.department,
            oa_total_budget=oa.total_budget,
            oa_version=oa.oa_version,
            status="draft",
            total_amount=Decimal("0"),
            purchase_type="询比价",
        )
        session.add(request)
        session.flush()

    total = request.total_amount or Decimal("0")
    written = 0
    for row in writable:
        matched = row["matched"]
        raw = row["raw"]
        quantity = Decimal(str(raw["quantity"]))
        unit_price = money(Decimal(str(matched["standard_price"])))
        line_amount = money(quantity * unit_price)
        confidence = (
            Decimal(str(row["match_confidence"]))
            if row.get("match_confidence")
            else Decimal("1.0000")
        )
        line = ProcurementRequestLine(
            material_code=matched["material_code"],
            material_name=matched["material_name"],
            specification=matched["specification"],
            unit=matched["unit"],
            quantity=quantity,
            unit_price=unit_price,
            line_amount=line_amount,
            raw_material_name=raw.get("material_name"),
            raw_specification=raw.get("specification"),
            raw_unit=raw.get("unit"),
            raw_quantity=quantity,
            raw_estimated_unit_price=(
                Decimal(str(raw["unit_price"])) if raw.get("unit_price") else None
            ),
            import_batch_id=batch.import_batch_id,
            match_confidence=confidence,
        )
        request.lines.append(line)
        session.flush()
        if raw.get("material_name") and raw["material_name"] != matched["material_name"]:
            session.add(
                CrossSystemDifference(
                    request_id=request.id,
                    request_line_id=line.id,
                    field_name="material_name",
                    source_system="EXCEL",
                    provided_value=str(raw["material_name"]),
                    authoritative_value=matched["material_name"],
                )
            )
        if raw.get("unit_price") and money(Decimal(str(raw["unit_price"]))) != unit_price:
            session.add(
                CrossSystemDifference(
                    request_id=request.id,
                    request_line_id=line.id,
                    field_name="unit_price",
                    source_system="EXCEL",
                    provided_value=str(raw["unit_price"]),
                    authoritative_value=str(unit_price),
                )
            )
        total += line_amount
        written += 1

    request.total_amount = money(total)
    ensure_purchase_method(request)
    batch.status = "confirmed"
    batch.success_rows = written
    batch.failed_rows = len(skipped)
    preview = dict(batch.preview_json or {})
    preview["confirmed_pr_no"] = request.request_no
    preview["skipped"] = skipped
    batch.preview_json = preview
    record_workflow_event(
        session,
        business_key=business_key or request.request_no,
        event_type="excel_import_confirmed",
        status="success",
        operator=task_id,
        detail={
            "import_batch_id": batch.import_batch_id,
            "pr_no": request.request_no,
            "written_rows": written,
            "skipped_rows": len(skipped),
        },
    )
    session.commit()
    request = get_request(session, request.id)
    return {
        "pr_no": request.request_no,
        "import_batch_id": batch.import_batch_id,
        "written_rows": written,
        "skipped_rows": len(skipped),
        "request": ProcurementRequestOut.model_validate(request).model_dump(mode="json"),
        "idempotent_replay": False,
    }


def list_export_candidates(
    session: Session,
    *,
    department: str | None = None,
    status: str | None = None,
    keyword: str | None = None,
    min_amount: Decimal | None = None,
    max_amount: Decimal | None = None,
    exportable_only: bool = False,
) -> list[dict[str, Any]]:
    rules = get_purchase_method_rules()
    statement = select(ProcurementRequest).options(selectinload(ProcurementRequest.lines))
    filters = []
    if department:
        filters.append(ProcurementRequest.oa_department == department)
    if status:
        filters.append(ProcurementRequest.status == status)
    if min_amount is not None:
        filters.append(ProcurementRequest.total_amount >= min_amount)
    if max_amount is not None:
        filters.append(ProcurementRequest.total_amount <= max_amount)
    if keyword:
        term = f"%{keyword}%"
        filters.append(
            or_(
                ProcurementRequest.request_no.ilike(term),
                ProcurementRequest.oa_applicant.ilike(term),
                ProcurementRequest.oa_title.ilike(term),
            )
        )
    if filters:
        statement = statement.where(*filters)
    requests = session.scalars(statement.order_by(ProcurementRequest.id.desc())).all()
    items = []
    for request in requests:
        ensure_purchase_method(request)
        validation = _validate_export_item(session, request)
        if exportable_only and validation["result"] == "blocked":
            continue
        items.append(
            {
                "pr_no": request.request_no,
                "department": request.oa_department,
                "applicant": request.oa_applicant,
                "title": request.oa_title,
                "status": request.status,
                "total_amount": str(money(request.total_amount)),
                "purchase_method_suggested": request.purchase_method_suggested,
                "purchase_method_confirmed": request.purchase_method_confirmed,
                "rule_version": request.rule_version or rules.get("version"),
                "export_status": request.export_status,
                "validation": validation,
            }
        )
    session.commit()
    return items


def _validate_export_item(session: Session, request: ProcurementRequest) -> dict[str, Any]:
    blockers: list[str] = []
    warnings: list[str] = []
    if request.status not in {"ready", "submitted", "validated"}:
        blockers.append("STATUS_NOT_EXPORTABLE")
    if not request.lines:
        blockers.append("EMPTY_LINES")
    active_codes = set(
        session.scalars(
            select(ERPMaterial.material_code).where(ERPMaterial.status == "active")
        )
    )
    invalid = [line.material_code for line in request.lines if line.material_code not in active_codes]
    if invalid:
        blockers.append("INVALID_ERP_MATERIAL")
    recomputed = money(sum((line.line_amount for line in request.lines), Decimal("0")))
    if recomputed != money(request.total_amount):
        warnings.append("AMOUNT_MISMATCH")
    if not request.purchase_method_confirmed:
        warnings.append("PURCHASE_METHOD_UNCONFIRMED")
    if blockers:
        result = "blocked"
    elif warnings:
        result = "review"
    else:
        result = "passed"
    return {
        "result": result,
        "blockers": blockers,
        "warnings": warnings,
        "invalid_material_codes": invalid,
        "recomputed_amount": str(recomputed),
    }


def batch_validate_requests(session: Session, pr_nos: list[str]) -> dict[str, Any]:
    rules = get_purchase_method_rules()
    results = []
    for pr_no in pr_nos:
        request = get_request_by_ref(session, pr_no)
        ensure_purchase_method(request)
        validation = _validate_export_item(session, request)
        results.append(
            {
                "pr_no": request.request_no,
                "total_amount": str(money(request.total_amount)),
                "purchase_method_suggested": request.purchase_method_suggested,
                "purchase_method_confirmed": request.purchase_method_confirmed,
                "rule_version": request.rule_version or rules.get("version"),
                "validation": validation,
            }
        )
    session.commit()
    summary = {
        "passed": sum(1 for item in results if item["validation"]["result"] == "passed"),
        "review": sum(1 for item in results if item["validation"]["result"] == "review"),
        "blocked": sum(1 for item in results if item["validation"]["result"] == "blocked"),
    }
    return {"rule_version": rules.get("version"), "summary": summary, "items": results}


def batch_export_requests(
    session: Session,
    *,
    pr_nos: list[str],
    template_version: str = "V3.2",
    filters: dict | None = None,
) -> dict[str, Any]:
    validation = batch_validate_requests(session, pr_nos)
    blocked = [
        item["pr_no"]
        for item in validation["items"]
        if item["validation"]["result"] == "blocked"
    ]
    if blocked:
        raise AppError(
            422,
            "EXPORT_BLOCKED",
            "One or more requests failed hard validation",
            {"blocked": blocked, "validation": validation},
        )

    rules = get_purchase_method_rules()
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    export_task_id = f"EXP-{uuid.uuid4().hex[:12].upper()}"
    stamp = datetime.now().strftime("%Y%m%d")
    filename = f"{datetime.now().year}年采购申请_{stamp}_{export_task_id[-6:]}.xlsx"
    file_path = EXPORTS_DIR / filename

    workbook = Workbook()
    workbook.remove(workbook.active)
    grouped: dict[str, list[ProcurementRequest]] = {}
    for pr_no in pr_nos:
        request = get_request_by_ref(session, pr_no)
        ensure_purchase_method(request)
        method = request.purchase_method_confirmed or request.purchase_method_suggested or "未确认"
        grouped.setdefault(method, []).append(request)
        request.export_status = "generated"

    for method, requests in grouped.items():
        sheet = workbook.create_sheet(title=str(method)[:31] or "导出")
        sheet.append(
            [
                "申请编号",
                "申请部门",
                "申请人",
                "标题",
                "状态",
                "金额",
                "建议采买方式",
                "确认采买方式",
                "规则版本",
            ]
        )
        for request in requests:
            sheet.append(
                [
                    request.request_no,
                    request.oa_department,
                    request.oa_applicant,
                    request.oa_title,
                    request.status,
                    float(request.total_amount),
                    request.purchase_method_suggested,
                    request.purchase_method_confirmed,
                    request.rule_version,
                ]
            )
    workbook.save(file_path)
    file_url = f"/api/v1/exports/{filename}"
    batch = ExportBatch(
        export_task_id=export_task_id,
        filters_json=filters or {},
        selected_prs=pr_nos,
        template_version=template_version,
        rule_version=str(rules.get("version", "v1")),
        file_path=str(file_path),
        file_url=file_url,
        status="generated",
    )
    session.add(batch)
    record_workflow_event(
        session,
        business_key=pr_nos[0] if pr_nos else export_task_id,
        event_type="batch_export",
        status="success",
        detail={"export_task_id": export_task_id, "pr_nos": pr_nos, "file_url": file_url},
    )
    session.commit()
    return {
        "export_task_id": export_task_id,
        "file_url": file_url,
        "filename": filename,
        "template_version": template_version,
        "rule_version": rules.get("version"),
        "selected_prs": pr_nos,
        "validation": validation,
    }


def patch_purchase_method(
    session: Session,
    pr_no: str,
    *,
    purchase_method_confirmed: str,
    task_id: str | None = None,
) -> dict[str, Any]:
    request = get_request_by_ref(session, pr_no)
    ensure_purchase_method(request)
    previous = request.purchase_method_confirmed
    request.purchase_method_confirmed = purchase_method_confirmed
    session.add(
        CrossSystemDifference(
            request_id=request.id,
            request_line_id=None,
            field_name="purchase_method_confirmed",
            source_system="USER",
            provided_value=previous,
            authoritative_value=purchase_method_confirmed,
        )
    )
    record_workflow_event(
        session,
        business_key=request.request_no,
        event_type="purchase_method_confirmed",
        status="success",
        operator=task_id,
        detail={
            "previous": previous,
            "confirmed": purchase_method_confirmed,
            "suggested": request.purchase_method_suggested,
        },
    )
    session.commit()
    return ProcurementRequestOut.model_validate(
        get_request(session, request.id)
    ).model_dump(mode="json")


ACTIVE_TASK_STATUSES = {
    "pending",
    "running",
    "wait_user",
    "WAIT_USER",
    "paused",
    "partial",
    "failed",
}


def list_active_agent_tasks(session: Session, limit: int = 20) -> list[dict[str, Any]]:
    tasks = session.scalars(
        select(AgentTask)
        .where(
            or_(
                AgentTask.status.in_(sorted(ACTIVE_TASK_STATUSES)),
                AgentTask.is_paused.is_(True),
            )
        )
        .order_by(AgentTask.id.desc())
        .limit(limit)
    ).all()
    return [_task_out(task) for task in tasks]


def _task_out(task: AgentTask) -> dict[str, Any]:
    return {
        "task_id": task.task_id,
        "business_key": task.business_key,
        "operation": task.operation,
        "status": task.status,
        "result": task.result or {},
        "current_route": task.current_route,
        "context_json": task.context_json or {},
        "is_paused": bool(task.is_paused),
        "created_at": task.created_at.isoformat() if task.created_at else None,
    }


def _latest_task(session: Session, task_id: str) -> AgentTask:
    task = session.scalar(
        select(AgentTask).where(AgentTask.task_id == task_id).order_by(AgentTask.id.desc())
    )
    if task is None:
        raise AppError(404, "TASK_NOT_FOUND", "Agent task not found")
    return task


def pause_agent_task(session: Session, task_id: str) -> dict[str, Any]:
    task = _latest_task(session, task_id)
    task.is_paused = True
    task.status = "paused"
    session.commit()
    return _task_out(task)


def resume_agent_task(session: Session, task_id: str) -> dict[str, Any]:
    task = _latest_task(session, task_id)
    task.is_paused = False
    if task.status in {"paused", "stopped"}:
        task.status = "running"
    session.commit()
    return _task_out(task)


def stop_agent_task(session: Session, task_id: str) -> dict[str, Any]:
    task = _latest_task(session, task_id)
    task.is_paused = False
    task.status = "stopped"
    session.commit()
    return _task_out(task)


def resolve_export_file(filename: str) -> Path:
    safe_name = Path(filename).name
    path = EXPORTS_DIR / safe_name
    if not path.exists() or not path.is_file():
        raise AppError(404, "EXPORT_NOT_FOUND", "Export file not found")
    return path
