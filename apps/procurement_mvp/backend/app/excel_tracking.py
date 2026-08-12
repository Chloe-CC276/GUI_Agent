"""Excel batch tracking snapshot for ERP PO Agent (fallback / audit, not SoT)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook


def tracking_dir() -> Path:
    root = Path(__file__).resolve().parents[2] / ".cache" / "erp_po_tracking"
    root.mkdir(parents=True, exist_ok=True)
    return root


def write_batch_excel(
    batch_id: str,
    *,
    operator: str | None,
    tasks: list[dict[str, Any]],
    lines: list[dict[str, Any]],
    audits: list[dict[str, Any]] | None = None,
) -> Path:
    path = tracking_dir() / f"{batch_id}.xlsx"
    wb = Workbook()
    batch_ws = wb.active
    batch_ws.title = "Batch"
    batch_ws.append(
        ["batch_id", "created_at", "operator", "total_count", "success", "failed", "wait_user"]
    )
    success = sum(1 for t in tasks if str(t.get("status", "")).upper() in {"SUCCESS", "PO_CREATED"})
    failed = sum(
        1
        for t in tasks
        if str(t.get("status", "")).upper() in {"FAILED", "DUPLICATE_BLOCKED"}
    )
    wait_user = sum(1 for t in tasks if str(t.get("status", "")).upper() == "WAIT_USER")
    batch_ws.append(
        [
            batch_id,
            tasks[0].get("created_at") if tasks else "",
            operator or "",
            len(tasks),
            success,
            failed,
            wait_user,
        ]
    )

    task_ws = wb.create_sheet("PO_Task")
    task_ws.append(
        [
            "task_id",
            "pr_no",
            "oa_apply_no",
            "supplier",
            "amount",
            "status",
            "retry_count",
            "po_no",
            "error_code",
            "source_snapshot_hash",
        ]
    )
    for task in tasks:
        task_ws.append(
            [
                task.get("task_id"),
                task.get("pr_no"),
                task.get("oa_apply_no"),
                task.get("supplier_name") or task.get("supplier_code"),
                task.get("total_amount"),
                task.get("status"),
                task.get("retry_count", 0),
                task.get("po_no"),
                task.get("error_code"),
                task.get("source_snapshot_hash"),
            ]
        )

    line_ws = wb.create_sheet("PO_Line")
    line_ws.append(
        [
            "pr_no",
            "line_no",
            "material_code",
            "material_name",
            "qty",
            "uom",
            "unit_price",
            "delivery_date",
        ]
    )
    for line in lines:
        line_ws.append(
            [
                line.get("pr_no"),
                line.get("line_no") or line.get("po_item_no"),
                line.get("material_code"),
                line.get("material_name"),
                line.get("quantity"),
                line.get("uom") or line.get("unit"),
                line.get("unit_price_tax") or line.get("unit_price"),
                line.get("delivery_date"),
            ]
        )

    audit_ws = wb.create_sheet("Audit")
    audit_ws.append(["step", "expected", "actual", "ref", "time", "result"])
    for item in audits or []:
        audit_ws.append(
            [
                item.get("step"),
                item.get("expected"),
                item.get("actual"),
                item.get("ref"),
                item.get("time"),
                item.get("result"),
            ]
        )

    wb.save(path)
    return path
