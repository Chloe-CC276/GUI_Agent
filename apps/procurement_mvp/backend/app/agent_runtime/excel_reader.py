"""Programmatic OA Excel → structured JSON (no GUI cell reading)."""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..errors import AppError

HEADER_ALIASES = {
    "title": {"title", "申请标题", "标题"},
    "applicant": {"applicant", "申请人"},
    "department": {"department", "申请部门", "部门"},
    "budget_project_name": {"budget_project_name", "预算项目名称", "预算项目"},
    "budget_project_code": {"budget_project_code", "预算项目编码"},
    "cost_center_code": {"cost_center_code", "成本中心"},
    "requested_method": {"requested_method", "建议采买方式", "采买方式"},
    "purchase_reason": {"purchase_reason", "采购原因"},
    "urgency_level": {"urgency_level", "紧急程度"},
    "expected_completion_date": {"expected_completion_date", "期望完成日期"},
    "remark": {"remark", "备注"},
    "item_name": {"item_name", "需求物资", "物资名称", "物料名称"},
    "specification": {"specification", "规格型号", "规格"},
    "quantity": {"quantity", "数量"},
    "estimated_unit_price": {"estimated_unit_price", "预估单价", "单价"},
    "total_budget": {"total_budget", "预算总额"},
    "application_no": {"application_no", "OA申请编号", "申请编号"},
}


def _norm(value: object) -> str:
    return str(value or "").strip()


def _money(value: object) -> Decimal:
    text = _norm(value).replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise AppError(422, "EXCEL_PARSE_FAILED", f"Invalid amount: {value}") from exc


def _map_headers(headers: list[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        name = _norm(header).lower()
        for field, aliases in HEADER_ALIASES.items():
            if name in {alias.lower() for alias in aliases}:
                mapping[field] = index
                break
    return mapping


def _cell(row: tuple, mapping: dict[str, int], field: str) -> object:
    index = mapping.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def list_excel_files(directory: str | Path) -> list[dict[str, str]]:
    root = Path(directory)
    if not root.exists() or not root.is_dir():
        raise AppError(400, "FOLDER_INVALID", "Directory does not exist", {"path": str(directory)})
    files = sorted(
        [path for path in root.iterdir() if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"}],
        key=lambda item: item.name.lower(),
    )
    return [{"name": path.name, "path": str(path.resolve())} for path in files]


def parse_oa_excel(
    source: bytes | str | Path,
    *,
    department_filter: str | None = None,
) -> dict[str, Any]:
    if isinstance(source, (str, Path)):
        workbook = load_workbook(filename=str(source), read_only=True, data_only=True)
    else:
        workbook = load_workbook(filename=BytesIO(source), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        raise AppError(422, "EXCEL_PARSE_FAILED", "Excel is empty")

    mapping = _map_headers(list(rows[0]))
    required = {"title", "department", "applicant", "item_name", "quantity", "estimated_unit_price"}
    missing = sorted(field for field in required if field not in mapping)
    if missing:
        raise AppError(422, "EXCEL_PARSE_FAILED", "Excel missing columns", {"missing": missing})

    # Group flat rows by application identity (title+department+applicant or application_no).
    groups: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for raw in rows[1:]:
        if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
            continue
        title = _norm(_cell(raw, mapping, "title"))
        department = _norm(_cell(raw, mapping, "department"))
        applicant = _norm(_cell(raw, mapping, "applicant"))
        app_no = _norm(_cell(raw, mapping, "application_no"))
        key = app_no or f"{department}|{applicant}|{title}"
        if key not in groups:
            groups[key] = {
                "application_no": app_no or None,
                "title": title,
                "department": department,
                "applicant": applicant,
                "budget_project_name": _norm(_cell(raw, mapping, "budget_project_name")),
                "budget_project_code": _norm(_cell(raw, mapping, "budget_project_code")) or None,
                "cost_center_code": _norm(_cell(raw, mapping, "cost_center_code")) or None,
                "requested_method": _norm(_cell(raw, mapping, "requested_method")) or None,
                "purchase_reason": _norm(_cell(raw, mapping, "purchase_reason")),
                "urgency_level": _norm(_cell(raw, mapping, "urgency_level")) or "NORMAL",
                "expected_completion_date": _norm(_cell(raw, mapping, "expected_completion_date")) or None,
                "remark": _norm(_cell(raw, mapping, "remark")) or None,
                "total_budget": _money(_cell(raw, mapping, "total_budget")),
                "lines": [],
            }
            order.append(key)
        item_name = _norm(_cell(raw, mapping, "item_name"))
        if not item_name:
            continue
        qty = _money(_cell(raw, mapping, "quantity"))
        price = _money(_cell(raw, mapping, "estimated_unit_price"))
        groups[key]["lines"].append(
            {
                "item_name": item_name,
                "specification": _norm(_cell(raw, mapping, "specification")),
                "quantity": qty,
                "estimated_unit_price": price,
                "line_amount": (qty * price).quantize(Decimal("0.01")),
            }
        )

    applications = [groups[key] for key in order]
    if department_filter:
        applications = [item for item in applications if item.get("department") == department_filter]
    if not applications:
        raise AppError(
            422,
            "EXCEL_NO_MATCH",
            "No purchase application matched the filter",
            {"department": department_filter},
        )

    # One application per import task (first match after filter).
    application = applications[0]
    if not application["lines"]:
        raise AppError(422, "EXCEL_PARSE_FAILED", "Application has no material lines")

    line_total = sum((line["line_amount"] for line in application["lines"]), Decimal("0"))
    # Line amounts are authoritative for import; header total is recalculated.
    application["total_budget"] = line_total if line_total > 0 else application["total_budget"]
    application["line_total"] = line_total
    application["matched_count"] = len(applications)
    return application


def validate_application_payload(application: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if not application.get("title"):
        errors.append("title is required")
    if not application.get("department"):
        errors.append("department is required")
    if not application.get("applicant"):
        errors.append("applicant is required")
    if not application.get("budget_project_name") and not application.get("budget_project_code"):
        errors.append("budget_project_name or budget_project_code is required")
    reason = str(application.get("purchase_reason") or "")
    if len(reason) < 10:
        errors.append("purchase_reason must be at least 10 characters")
    if not application.get("urgency_level"):
        errors.append("urgency_level is required")
    lines = application.get("lines") or []
    if not lines:
        errors.append("at least one material line is required")
    for index, line in enumerate(lines):
        if not line.get("item_name"):
            errors.append(f"line[{index}].item_name is required")
        if Decimal(str(line.get("quantity") or 0)) <= 0:
            errors.append(f"line[{index}].quantity must be > 0")
        if Decimal(str(line.get("estimated_unit_price") or 0)) < 0:
            errors.append(f"line[{index}].estimated_unit_price must be >= 0")
    total = Decimal(str(application.get("total_budget") or 0))
    line_total = Decimal(str(application.get("line_total") or 0))
    if total <= 0:
        errors.append("total_budget must be > 0")
    elif line_total > 0 and abs(total - line_total) > Decimal("0.05"):
        errors.append(f"total_budget {total} != line sum {line_total}")
    return errors
